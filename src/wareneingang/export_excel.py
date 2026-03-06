# src/wareneingang/export_excel.py
from typing import List, Dict

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows

FILL_OK = PatternFill("solid", fgColor="C6EFCE")
FILL_PARTIAL = PatternFill("solid", fgColor="FFEB9C")
FILL_PARKED = PatternFill("solid", fgColor="FFC7CE")


def export_status_xlsx(status_rows: List[Dict], out_path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Status"

    df = pd.DataFrame(status_rows)
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    for cell in ws[1]:
        cell.font = Font(bold=True)

    headers = [c.value for c in ws[1]]
    status_col = headers.index("status") + 1

    for row in range(2, ws.max_row + 1):
        st = ws.cell(row=row, column=status_col).value
        if st == "OK":
            fill = FILL_OK
        elif st == "PARTIAL":
            fill = FILL_PARTIAL
        else:
            fill = FILL_PARKED
        for col in range(1, ws.max_column + 1):
            ws.cell(row=row, column=col).fill = fill

    wb.save(out_path)